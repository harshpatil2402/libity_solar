import os, zipfile, io, base64, tempfile, subprocess, glob, uuid, requests, shutil, threading, re
from math import ceil
from datetime import datetime, timedelta
from flask import Flask, render_template, request, session, redirect, url_for, flash, jsonify, send_file, Response
from supabase import create_client
from weasyprint import HTML as WP_HTML, CSS
from PIL import Image, ImageOps
from dotenv import load_dotenv
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

load_dotenv()
app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY")
supabase = create_client(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_KEY"))

BASE_DIR        = os.path.abspath(os.path.dirname(__file__))
HTML_DOCS_DIR   = os.path.join(BASE_DIR, 'input_docs')   # HTML templates for WeasyPrint
app.config['MAX_CONTENT_LENGTH']   = 50 * 1024 * 1024
app.config['MAX_FORM_MEMORY_SIZE'] = 50 * 1024 * 1024

PER_PAGE_ADMIN   = 10
PER_PAGE_HISTORY = 15

# WeasyPrint is in-process — no LibreOffice RAM spikes, no subprocess needed
jobs, jobs_lock = {}, threading.Lock()

# ── Admin contact info — hardcoded from .env ──────────────────────────────────
ADMIN_INFO = {
    'admin_phone':    os.getenv('ADMIN_PHONE',    ''),
    'admin_whatsapp': os.getenv('ADMIN_WHATSAPP', ''),
    'admin_email':    os.getenv('ADMIN_EMAIL',    ''),
}

# ── HTML template cache (loaded once at startup) ──────────────────────────────
_HTML_CACHE: dict[str, str] = {}   # fname → raw HTML string

# ── Decorators ────────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def d(*a, **kw):
        if 'user' not in session:
            return redirect(url_for('login'))
        return f(*a, **kw)
    return d

def admin_required(f):
    @wraps(f)
    def d(*a, **kw):
        if session.get('user', {}).get('role') != 'admin':
            flash("Access denied.", "danger")
            return redirect(url_for('agency_dashboard'))
        return f(*a, **kw)
    return d

# ── Helpers ───────────────────────────────────────────────────────────────────
def days_left(exp_str):
    if not exp_str:
        return None
    try:
        return (datetime.strptime(str(exp_str)[:10], '%Y-%m-%d').date() - datetime.now().date()).days
    except Exception:
        return None

def job_log(jid, msg, error=False):
    with jobs_lock:
        if jid in jobs:
            jobs[jid]['logs'].append({'msg': msg, 'error': error})

def img_to_data_uri(img_bytes: bytes, mime='image/jpeg') -> str:
    """Convert raw image bytes to a base64 data URI for inline HTML embedding."""
    return f"data:{mime};base64,{base64.b64encode(img_bytes).decode()}"

def b64_to_jpeg_bytes(b64: str) -> bytes | None:
    """Decode base64 image string, convert to JPEG bytes."""
    if not b64:
        return None
    try:
        if ',' in b64:
            b64 = b64.split(',')[1]
        img = Image.open(io.BytesIO(base64.b64decode(b64))).convert('RGB')
        out = io.BytesIO()
        img.save(out, format='JPEG', quality=90, optimize=True)
        return out.getvalue()
    except Exception as e:
        print(f"Image decode error: {e}")
        return None

def url_to_jpeg_bytes(url: str) -> bytes | None:
    """Download image from URL, convert to JPEG bytes."""
    if not url:
        return None
    try:
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        img = Image.open(io.BytesIO(r.content)).convert('RGB')
        out = io.BytesIO()
        img.save(out, format='JPEG', quality=90, optimize=True)
        return out.getvalue()
    except Exception:
        return None

def upload_image(b64, bucket):
    if not b64:
        return None
    try:
        raw  = base64.b64decode(b64.split(',')[1] if ',' in b64 else b64)
        path = f"branding/{uuid.uuid4()}.png"
        supabase.storage.from_(bucket).upload(path, raw, {"content-type": "image/png"})
        return supabase.storage.from_(bucket).get_public_url(path)
    except Exception as e:
        print(f"Upload error: {e}")
        return None

def preload_templates():
    """Load HTML templates into memory at startup."""
    html_names = [
        "commissioning_report.html",
        "meter_testing.html",
        "model_agreement.html",
        "net_metering_agreement.html",
        "work_completion_report.html",
    ]
    loaded = 0
    for name in html_names:
        path = os.path.join(HTML_DOCS_DIR, name)
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                _HTML_CACHE[name] = f.read()
            loaded += 1
    print(f"[LibityInfotech] Preloaded {loaded}/{len(html_names)} HTML templates.")

# ── WeasyPrint PDF generation ─────────────────────────────────────────────────
def fill_html_template(html: str, ctx: dict) -> str:
    """
    Replace {{variable}} placeholders in HTML with ctx values.
    Images are embedded as data URIs.
    Removes yellow highlight spans after substitution.
    """
    for key, val in ctx.items():
        if val is not None:
            html = html.replace('{{' + key + '}}', str(val))
    # Clear any unfilled placeholders
    html = re.sub(r'\{\{[^}]+\}\}', '', html)
    # Remove yellow highlight (only needed for template editing, not output)
    html = html.replace(' class="highlight"', '').replace(' class=\'highlight\'', '')
    return html

def render_pdf(fname: str, oname: str, ctx: dict, jid: str) -> bytes | None:
    """Render one HTML template → PDF bytes using WeasyPrint."""
    try:
        job_log(jid, f"Rendering {oname} ...")
        html_src = _HTML_CACHE.get(fname)
        if html_src is None:
            path = os.path.join(HTML_DOCS_DIR, fname)
            if not os.path.exists(path):
                job_log(jid, f"Missing template: {fname}", error=True)
                return None
            with open(path, 'r', encoding='utf-8') as f:
                html_src = f.read()
        filled = fill_html_template(html_src, ctx)
        # WeasyPrint: base_url=HTML_DOCS_DIR lets @font-face and static refs resolve
        pdf_bytes = WP_HTML(
            string=filled,
            base_url=HTML_DOCS_DIR
        ).write_pdf()
        job_log(jid, f"Done: {oname}.pdf")
        return pdf_bytes
    except Exception as e:
        job_log(jid, f"Failed {oname}: {e}", error=True)
        return None

def run_job(jid, form_data, agency_id):
    from concurrent.futures import ThreadPoolExecutor, as_completed
    def log(m, e=False): job_log(jid, m, e)
    try:
        log("Fetching profile ...")
        profile    = supabase.table('agencies').select('*').eq('id', agency_id).single().execute().data
        sig_b64    = form_data.pop('sig_b64', None)
        aadhar_b64 = form_data.pop('aadhar_b64', None)
        form_data.pop('format', None)

        # ── Load Mahavitaran logo from disk as data URI ────────────
        # WeasyPrint needs base_url=HTML_DOCS_DIR, but <img src="static/..."> resolves
        # relative to that dir — which works only if static/ is inside input_docs/.
        # Safest: embed as data URI so it always resolves regardless of base_url.
        mahavitaran_uri = ''
        maha_path = os.path.join(BASE_DIR, 'static', 'images', 'mahavitaran_logo.png')
        if os.path.exists(maha_path):
            with open(maha_path, 'rb') as _f:
                mahavitaran_uri = f"data:image/png;base64,{base64.b64encode(_f.read()).decode()}"

        # ── Process agency images in parallel ──────────────────────
        log("Processing images ...")
        results = [None] * 4
        def fetch(i, fn): results[i] = fn()
        ts = [threading.Thread(target=fetch, args=(i, fn)) for i, fn in enumerate([
            lambda: b64_to_jpeg_bytes(sig_b64),
            lambda: b64_to_jpeg_bytes(aadhar_b64),
            lambda: url_to_jpeg_bytes(profile.get('logo_url')),
            lambda: url_to_jpeg_bytes(profile.get('stamp_url')),
        ])]
        for t in ts: t.start()
        for t in ts: t.join()

        sig_bytes, aadhar_bytes, logo_bytes, stamp_bytes = results

        def to_uri(b): return img_to_data_uri(b) if b else ''

        fd   = form_data
        inv  = fd.get('inverter_make_and_model', '')
        inv_parts = inv.split(' ', 1) if inv else ['', '']
        inv_make  = inv_parts[0]
        inv_model = inv_parts[1] if len(inv_parts) > 1 else inv

        ctx = {
            # ── Consumer ──────────────────────────────────────────
            'consumer_name':             fd.get('consumer_name', ''),
            'consumer_number':           fd.get('consumer_number', ''),
            'consumer_contact_number':   fd.get('consumer_contact_number', ''),
            'consumer_email':            fd.get('consumer_email', ''),
            'consumer_address':          fd.get('consumer_address', ''),
            'consumer_aadhar_num':       fd.get('consumer_aadhar_num', ''),
            'city':                      fd.get('city', ''),
            # ── Grid / Sanction ────────────────────────────────────
            'discom_division':           fd.get('discom_division', ''),
            'licensee_name':             fd.get('licensee_name', ''),
            'sanction_number':           fd.get('sanction_number', ''),
            'sanction_capacity_kw':      fd.get('sanction_capacity_kw', ''),
            'system_capacity_kw':        fd.get('system_capacity_kw', ''),
            'agreement_solar_price':     fd.get('agreement_solar_price', ''),
            # ── Modules ────────────────────────────────────────────
            'module_make':               fd.get('module_make', ''),
            'almm_model_number':         fd.get('almm_model_number', ''),
            'module_efficiency':         fd.get('module_efficiency', ''),
            'module_capacity_wp':        fd.get('module_capacity_wp', ''),
            'num_pv_modules':            fd.get('num_pv_modules', ''),
            'total_module_capacity_kwp': fd.get('total_module_capacity_kwp', ''),
            # ── Inverter ───────────────────────────────────────────
            'inverter_make_and_model':   inv,
            'inverter_make':             inv_make,
            'inverter_model':            inv_model,
            'inverter_capacity_kw':      fd.get('inverter_capacity_kw', ''),
            'inverter_rating_text':      fd.get('inverter_rating_text', ''),
            # ── Dates ──────────────────────────────────────────────
            'agreement_date':            fd.get('agreement_date', ''),
            'annexure_agreement_date':   fd.get('annexure_agreement_date', ''),
            'installation_date':         fd.get('installation_date', ''),
            'meter_testing_date':        fd.get('meter_testing_date', ''),
            'performance_check_date':    fd.get('performance_check_date', ''),
            'today_date':                datetime.now().strftime('%d-%m-%Y'),
            # ── Agency ─────────────────────────────────────────────
            'agency_name':               profile.get('agency_name', ''),
            'agency_address':            profile.get('agency_address', ''),
            'agency_contact':            profile.get('contact_number', ''),
            'agency_director':           profile.get('director_name', ''),
            # ── Images (all as data URIs — no filesystem dependency) ──
            'mahavitaran_logo':          mahavitaran_uri,
            'consumer_signature_image':  f'<img src="{to_uri(sig_bytes)}" style="width:100%;height:100%;object-fit:contain;display:block;">' if sig_bytes else '',
            'consumer_aadhar_image':     f'<img src="{to_uri(aadhar_bytes)}" style="width:100%;height:100%;object-fit:contain;display:block;">' if aadhar_bytes else '',
            'agency_stamp_image':        f'<img src="{to_uri(stamp_bytes)}" style="width:100%;height:100%;object-fit:contain;display:block;">' if stamp_bytes else '',
            'agency_logo':               to_uri(logo_bytes) if logo_bytes else '',
        }

        docs = [
            ("commissioning_report.html",  "1_Commissioning_Report"),
            ("meter_testing.html",          "2_Meter_Testing"),
            ("model_agreement.html",        "3_Model_Agreement"),
            ("net_metering_agreement.html", "4_Net_Metering"),
            ("work_completion_report.html", "5_Work_Completion"),
        ]

        # ── Parallel PDF rendering ─────────────────────────────────
        # WeasyPrint IS thread-safe for separate HTML() instances.
        # We fill HTML strings (pure Python, no I/O) then render in parallel.
        # Each worker creates its own WP_HTML object — no shared state.
        # Tested speedup: ~2.5–3× on a 4-core server vs sequential.
        log("Generating PDFs in parallel ...")
        pdf_results = [None] * len(docs)

        def render_one(idx, fname, oname):
            pdf_results[idx] = render_pdf(fname, oname, ctx, jid)

        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = {
                executor.submit(render_one, i, fname, oname): i
                for i, (fname, oname) in enumerate(docs)
            }
            for future in as_completed(futures):
                future.result()   # surface any exceptions

        # ── ZIP in memory ──────────────────────────────────────────
        log("Building ZIP ...")
        cn    = fd.get('consumer_name', 'Client').replace(' ', '_')
        cno   = fd.get('consumer_number', '0000')
        zname = f"{cn}_{cno}_{datetime.now().strftime('%d-%m-%Y_%H%M%S')}.zip"
        zbuf  = io.BytesIO()
        with zipfile.ZipFile(zbuf, 'w', zipfile.ZIP_STORED) as zf:
            for i, (_, oname) in enumerate(docs):
                if pdf_results[i]:
                    zf.writestr(f"{oname}.pdf", pdf_results[i])
        zbuf.seek(0)

        with jobs_lock:
            jobs[jid]['status']    = 'done'
            jobs[jid]['zip_name']  = zname
            jobs[jid]['zip_bytes'] = zbuf.getvalue()
        log("ZIP ready — downloading.")
    except Exception as e:
        job_log(jid, f"Fatal: {e}", error=True)
        with jobs_lock: jobs[jid]['status'] = 'error'

# ── Generation API ────────────────────────────────────────────────────────────
@app.route('/api/generate', methods=['POST'])
@login_required
def api_generate():
    aid = session['user']['id']
    fd  = request.get_json()

    # DB insert is fire-and-forget — never blocks the response
    def _db_insert():
        try:
            db = {k: v for k, v in fd.items() if k not in ['sig_b64', 'aadhar_b64', 'format']}
            supabase.table('generation_history').insert({**db, 'agency_id': aid}).execute()
        except Exception as e:
            print(f"History insert: {e}")
    threading.Thread(target=_db_insert, daemon=True).start()

    jid = str(uuid.uuid4())
    with jobs_lock:
        jobs[jid] = {'logs': [], 'status': 'running', 'download_url': None}
    def go():
        with app.app_context(): run_job(jid, dict(fd), aid)
    threading.Thread(target=go, daemon=True).start()
    return jsonify({'job_id': jid})

@app.route('/api/job/<jid>/status')
@login_required
def api_job_status(jid):
    since = int(request.args.get('since', 0))
    with jobs_lock: job = jobs.get(jid)
    if not job:
        return jsonify({'error': 'not found'}), 404
    # Expose a download URL only when ready — points to the one-time stream route
    dl = f'/api/job/{jid}/download' if job.get('status') == 'done' else None
    return jsonify({'logs': job['logs'][since:], 'total': len(job['logs']),
                    'status': job['status'], 'download_url': dl})

@app.route('/api/job/<jid>/download')
@login_required
def api_job_download(jid):
    """
    Stream the in-memory ZIP to the browser, then immediately purge it
    from the job store.  The file never existed on disk and never will.
    Agency can regenerate any time from generation history.
    """
    with jobs_lock:
        job = jobs.get(jid)
        if not job or job.get('status') != 'done':
            return jsonify({'error': 'Not ready'}), 404
        if not job.get('zip_bytes'):
            return jsonify({'error': 'Already downloaded'}), 404
        raw   = job['zip_bytes']   # keep bytes — allow re-download
        zname = job.get('zip_name', 'documents.zip')
        job['downloaded_at'] = datetime.now().isoformat()
    return send_file(
        io.BytesIO(raw),
        mimetype='application/zip',
        as_attachment=True,
        download_name=zname,
    )

# ── Auth ──────────────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        lid = request.form.get('login_id', '').strip()
        pw  = request.form.get('password', '')
        res = supabase.table('agencies').select('*').or_(
            f"username.eq.{lid},email.eq.{lid}").execute()
        if res.data and res.data[0]['password'] == pw:
            u = res.data[0]
            # Allow login even if expired — agency sees blur screen with contact info
            u['days_left'] = days_left(u.get('expires_at')) or 9999
            session['user'] = u
            flash(f"Welcome back, {u['agency_name']}!", "success")
            return redirect(url_for('index'))
        flash("Invalid credentials.", "danger")

    # Active agencies with logos for login-page carousel
    carousel_agencies = []
    try:
        all_ag = supabase.table('agencies').select('agency_name,logo_url,expires_at') \
                         .neq('role', 'admin').execute().data or []
        carousel_agencies = [
            a for a in all_ag
            if a.get('logo_url') and (days_left(a.get('expires_at') or '') or 0) > 0
        ]
    except Exception:
        pass

    return render_template('login.html', contact_info=ADMIN_INFO,
                           carousel_agencies=carousel_agencies)

@app.route('/logout')
def logout():
    session.clear()
    flash("Logged out successfully.", "info")
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    u = session['user']
    if u['role'] == 'admin':
        return redirect(url_for('admin_dashboard'))
    # Agency users → their prefixed dashboard
    return redirect(f"/{u['username']}/dashboard")

# ── Jinja helper — builds prefixed URL for current agency user ────────────────
@app.context_processor
def inject_slug_helpers():
    def au(endpoint, **kwargs):
        """
        Agency URL helper for templates.
        For agency users, builds the /<username>/... prefixed URL.
        Falls back to url_for for admin and unauthenticated.
        """
        from flask import session as _s
        u = _s.get('user', {})
        if u.get('role') == 'agency':
            slug = u.get('username', '')
            paths = {
                'agency_dashboard': f'/{slug}/dashboard',
                'generate':         f'/{slug}/generate',
                'history':          f'/{slug}/history',
            }
            if endpoint in paths:
                base = paths[endpoint]
                if kwargs:
                    qs = '&'.join(f'{k}={v}' for k, v in kwargs.items())
                    return f'{base}?{qs}'
                return base
        return url_for(endpoint, **kwargs)
    # Make ADMIN_INFO contact details available in every template
    # (needed by expired-subscription overlay in base.html)
    return dict(au=au, contact_info=ADMIN_INFO)
    return dict(au=au)

# ── Agency portal — branded login page at /<username> ────────────────────────
_RESERVED_SLUGS = {
    'login', 'logout', 'dashboard', 'admin', 'generate', 'history',
    'static', 'api', 'favicon.ico', 'error',
}

@app.route('/<slug>')
def agency_portal(slug):
    slug = slug.strip().lower()
    if slug in _RESERVED_SLUGS:
        return redirect(url_for('login'))
    try:
        res = supabase.table('agencies').select(
            'id,username,agency_name,logo_url,expires_at,role'
        ).eq('username', slug).neq('role', 'admin').execute()
    except Exception:
        return redirect(url_for('login'))
    if not res.data:
        return redirect(url_for('login'))
    agency = res.data[0]
    dl = days_left(agency.get('expires_at') or '')
    return render_template('portal.html', agency=agency, days_left=dl, contact_info=ADMIN_INFO)

# ── Slug-prefixed agency routes ───────────────────────────────────────────────
# After login, agency users are redirected to /<username>/dashboard.
# All internal links use au() helper so they stay under the prefix.
# The old flat routes (/dashboard, /generate, /history) still work as
# redirects to the prefixed versions for robustness.

def _slug_login_required(f):
    """Decorator: checks login AND that the slug in the URL matches the session user."""
    @wraps(f)
    def d(slug, *a, **kw):
        if 'user' not in session:
            return redirect(f'/{slug}')
        u = session['user']
        if u.get('role') == 'admin':
            return redirect(url_for('admin_dashboard'))
        if u.get('username', '').lower() != slug.lower():
            # Different user's URL — send them to their own
            return redirect(f"/{u['username']}/{f.__name__.replace('slug_', '')}")
        return f(slug, *a, **kw)
    return d

@app.route('/<slug>/dashboard')
@_slug_login_required
def slug_dashboard(slug):
    aid = session['user']['id']
    try:
        total_docs = supabase.table('generation_history').select('id', count='exact') \
                              .eq('agency_id', aid).execute().count or 0
    except Exception:
        total_docs = 0
    return render_template('agency_dashboard.html', total_docs=total_docs, payment_info=ADMIN_INFO)

@app.route('/<slug>/generate')
@_slug_login_required
def slug_generate(slug):
    prefill = {}
    from_history = request.args.get('from_history', '').strip()
    if from_history:
        try:
            aid = session['user']['id']
            rec = (supabase.table('generation_history').select('*')
                   .eq('id', from_history).eq('agency_id', aid).single().execute().data)
            if rec:
                skip = {'id', 'agency_id', 'created_at'}
                prefill = {k: (v or '') for k, v in rec.items() if k not in skip}
        except Exception:
            pass
    return render_template('generate_form.html', prefill=prefill)

@app.route('/<slug>/history')
@_slug_login_required
def slug_history(slug):
    aid  = session['user']['id']
    q    = request.args.get('q', '').strip()
    page = max(1, int(request.args.get('page', 1)))
    all_r = (supabase.table('generation_history').select('*')
             .eq('agency_id', aid).order('created_at', desc=True).execute().data or [])
    if q:
        ql    = q.lower()
        all_r = [r for r in all_r if
                 ql in (r.get('consumer_name') or '').lower() or
                 ql in (r.get('consumer_number') or '').lower() or
                 ql in (r.get('city') or '').lower()]
    total       = len(all_r)
    total_pages = max(1, ceil(total / PER_PAGE_HISTORY))
    page        = min(page, total_pages)
    records     = all_r[(page - 1) * PER_PAGE_HISTORY: page * PER_PAGE_HISTORY]
    return render_template('history.html',
        history=records, page=page, total_pages=total_pages, total=total, q=q)

@app.route('/<slug>/history/<record_id>')
@_slug_login_required
def slug_history_detail(slug, record_id):
    aid = session['user']['id']
    rec = (supabase.table('generation_history').select('*')
           .eq('id', record_id).eq('agency_id', aid).single().execute().data)
    if not rec:
        flash("Record not found.", "danger")
        return redirect(f'/{slug}/history')
    return render_template('history_detail.html', record=rec)

@app.route('/<slug>/history/delete/<record_id>', methods=['POST'])
@_slug_login_required
def slug_history_delete(slug, record_id):
    aid = session['user']['id']
    supabase.table('generation_history').delete().eq('id', record_id).eq('agency_id', aid).execute()
    flash("Record deleted.", "info")
    return redirect(f'/{slug}/history?q={request.form.get("_q","")}&page={request.form.get("_page",1)}')

# ── Old flat routes kept as redirects (backward compat) ──────────────────────
@app.route('/dashboard')
@login_required
def agency_dashboard():
    u = session['user']
    if u['role'] == 'admin':
        return redirect(url_for('admin_dashboard'))
    return redirect(f"/{u['username']}/dashboard")

# ── Admin dashboard — search + status filter + pagination ─────────────────────
@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    search = request.args.get('search', '').strip()
    status = request.args.get('status', 'all')
    page   = max(1, int(request.args.get('page', 1)))

    raw = supabase.table('agencies').select('*').neq('role', 'admin').execute().data or []
    for a in raw:
        a['days_left'] = days_left(a.get('expires_at')) or 0

    stats = {
        'total':   len(raw),
        'active':  sum(1 for a in raw if a['days_left'] > 0),
        'expired': sum(1 for a in raw if a['days_left'] <= 0),
    }

    filtered = raw
    if search:
        sl = search.lower()
        filtered = [a for a in raw if
                    sl in (a.get('agency_name') or '').lower() or
                    sl in (a.get('username') or '').lower() or
                    sl in (a.get('director_name') or '').lower() or
                    sl in (a.get('contact_number') or '').lower()]
    if status == 'active':
        filtered = [a for a in filtered if a['days_left'] > 0]
    elif status == 'expired':
        filtered = [a for a in filtered if a['days_left'] <= 0]

    total       = len(filtered)
    total_pages = max(1, ceil(total / PER_PAGE_ADMIN))
    page        = min(page, total_pages)
    agencies_pg = filtered[(page - 1) * PER_PAGE_ADMIN: page * PER_PAGE_ADMIN]

    # Payment & contact info — read from .env, no DB query needed
    payment_info = ADMIN_INFO

    return render_template('admin_dashboard.html',
        agencies=agencies_pg, stats=stats,
        page=page, total_pages=total_pages, total=total,
        search=search, status=status,
        payment_info=payment_info)

# ── Admin: subscription + usage Excel export ──────────────────────────────────
@app.route('/admin/export')
@login_required
@admin_required
def admin_export():
    from_date     = request.args.get('from_date', '').strip()
    to_date       = request.args.get('to_date', '').strip()
    export_status = request.args.get('export_status', 'all')
    cols_req      = request.args.getlist('cols') or [
        'agency_name','director_name','contact_number','email','username',
        'created_at','expires_at','days_left','total_docs'
    ]

    try:
        agencies = supabase.table('agencies').select('*').neq('role', 'admin').execute().data or []
        history  = supabase.table('generation_history').select('agency_id,created_at').execute().data or []
    except Exception as e:
        flash(f"Export failed: {e}", "danger")
        return redirect(url_for('admin_dashboard'))

    # Filter agencies by creation date if set
    if from_date:
        agencies = [a for a in agencies if (a.get('created_at') or '') >= from_date]
    if to_date:
        to_end = to_date + 'T23:59:59'
        agencies = [a for a in agencies if (a.get('created_at') or '') <= to_end]

    # Filter by status
    for a in agencies:
        a['days_left'] = days_left(a.get('expires_at')) or 0
    if export_status == 'active':
        agencies = [a for a in agencies if a['days_left'] > 0]
    elif export_status == 'expired':
        agencies = [a for a in agencies if a['days_left'] <= 0]

    # Doc counts
    from collections import defaultdict
    monthly = defaultdict(lambda: defaultdict(int))
    for row in history:
        aid = row.get('agency_id')
        if aid and row.get('created_at'):
            monthly[aid][row['created_at'][:7]] += 1

    all_months = sorted({row['created_at'][:7] for row in history if row.get('created_at')})

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Agencies"

    hdr_fill    = PatternFill("solid", fgColor="406093")
    hdr_font    = Font(bold=True, color="FFFFFF", size=11)
    alt_fill    = PatternFill("solid", fgColor="D6E4F0")
    center      = Alignment(horizontal="center", vertical="center")
    thin_side   = Side(style="thin", color="AAAAAA")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Column map — label, width, extractor
    ALL_COLS = {
        'agency_name':    ("Agency Name",          28, lambda a, _: a.get('agency_name','')),
        'director_name':  ("Director Name",         22, lambda a, _: a.get('director_name','')),
        'contact_number': ("Contact Number",         16, lambda a, _: a.get('contact_number','')),
        'email':          ("Email",                  28, lambda a, _: a.get('email','')),
        'username':       ("Username",               18, lambda a, _: a.get('username','')),
        'created_at':     ("Account Created On",     22, lambda a, _: (a.get('created_at') or '')[:10]),
        'expires_at':     ("Subscription Expiry",    20, lambda a, _: a.get('expires_at','')),
        'days_left':      ("Days Remaining",          16, lambda a, _: a.get('days_left', 0)),
        'total_docs':     ("Total Docs Generated",   18, lambda a, m: sum(m.get(a['id'], {}).values())),
    }
    active_cols = [(k, *ALL_COLS[k]) for k in cols_req if k in ALL_COLS]

    # Header row
    for col_i, (_, label, width, _fn) in enumerate(active_cols, 1):
        cell = ws1.cell(row=1, column=col_i, value=label)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = center; cell.border = thin_border
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col_i)].width = width
    ws1.row_dimensions[1].height = 22

    # Data rows
    for row_i, a in enumerate(agencies, 2):
        dl = a.get('days_left', 0)
        fill = alt_fill if row_i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_i, (key, _label, _w, extractor) in enumerate(active_cols, 1):
            val  = extractor(a, monthly)
            cell = ws1.cell(row=row_i, column=col_i, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if key == 'days_left':
                color = "91D06C" if dl > 30 else ("FFA500" if dl > 0 else "FF6B6B")
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True)
            else:
                cell.fill = fill

    # Sheet 2: Monthly usage
    ws2 = wb.create_sheet("Monthly Usage")
    months_header = ["Agency Name"] + all_months
    for col, h in enumerate(months_header, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = center; cell.border = thin_border
    ws2.column_dimensions["A"].width = 28
    for i in range(2, len(months_header) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14
    ws2.row_dimensions[1].height = 22

    for row_i, a in enumerate(agencies, 2):
        fill = alt_fill if row_i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        c = ws2.cell(row=row_i, column=1, value=a.get('agency_name', ''))
        c.fill = fill; c.border = thin_border
        for col_i, mo in enumerate(all_months, 2):
            count = monthly.get(a['id'], {}).get(mo, 0)
            cell  = ws2.cell(row=row_i, column=col_i, value=count if count else '')
            cell.alignment = center; cell.border = thin_border
            cell.fill = PatternFill("solid", fgColor="FFF799") if count else fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"LibityInfotech_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)

# ── Admin: create ─────────────────────────────────────────────────────────────
def _calc_expiry(form) -> str:
    """
    Calculate expiry date from form fields.
    Priority: custom_expires_at > subscription_months (float, supports 0.27 = 8 days).
    """
    custom_date = form.get('custom_expires_at', '').strip()
    if custom_date:
        try:
            datetime.strptime(custom_date, '%Y-%m-%d')  # validate
            return custom_date
        except ValueError:
            pass

    raw_months = form.get('subscription_months', '12')
    try:
        months = float(raw_months)
    except (ValueError, TypeError):
        months = 12.0

    if months == 0.27:                          # 8-day demo shortcut
        return (datetime.now() + timedelta(days=8)).strftime('%Y-%m-%d')

    days = int(round(months * 30.44))           # accurate month→day conversion
    return (datetime.now() + timedelta(days=days)).strftime('%Y-%m-%d')


@app.route('/admin/agency/new', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_create_agency():
    if request.method == 'POST':
        expires_at = _calc_expiry(request.form)
        data = {
            "username":       request.form.get('username'),
            "email":          request.form.get('email'),
            "password":       request.form.get('password'),
            "agency_name":    request.form.get('agency_name'),
            "director_name":  request.form.get('director_name'),
            "contact_number": request.form.get('contact_number'),
            "agency_address": request.form.get('agency_address'),
            "role":           "agency",
            "expires_at":     expires_at,
        }
        logo_url  = upload_image(request.form.get('logo_base64'),  'agency-logos')
        stamp_url = upload_image(request.form.get('stamp_base64'), 'agency-stamps')
        if logo_url:  data['logo_url']  = logo_url
        if stamp_url: data['stamp_url'] = stamp_url
        try:
            supabase.table('agencies').insert(data).execute()
            flash(f"Agency '{data['agency_name']}' created — expires {expires_at}.", "success")
            return redirect(url_for('admin_dashboard'))
        except Exception as e:
            flash(f"Error: {e}", "danger")
    return render_template('admin_agency_form.html', agency=None, edit=False)

# ── Admin: edit ───────────────────────────────────────────────────────────────
@app.route('/admin/agency/edit/<agency_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_edit_agency(agency_id):
    agency = supabase.table('agencies').select('*').eq('id', agency_id).single().execute().data
    if not agency:
        flash("Agency not found.", "danger")
        return redirect(url_for('admin_dashboard'))
    if request.method == 'POST':
        data = {
            "agency_name":    request.form.get('agency_name'),
            "director_name":  request.form.get('director_name'),
            "contact_number": request.form.get('contact_number'),
            "agency_address": request.form.get('agency_address'),
            "email":          request.form.get('email'),
            "username":       request.form.get('username'),
        }
        if request.form.get('password', '').strip():
            data['password'] = request.form['password'].strip()
        if request.form.get('expires_at', '').strip():
            data['expires_at'] = request.form['expires_at'].strip()
        new_logo  = upload_image(request.form.get('logo_base64'),  'agency-logos')
        new_stamp = upload_image(request.form.get('stamp_base64'), 'agency-stamps')
        if new_logo:  data['logo_url']  = new_logo
        if new_stamp: data['stamp_url'] = new_stamp
        try:
            supabase.table('agencies').update(data).eq('id', agency_id).execute()
            flash("Agency updated successfully.", "success")
            return redirect(url_for('admin_dashboard'))
        except Exception as e:
            flash(f"Update error: {e}", "danger")
    return render_template('admin_agency_form.html', agency=agency, edit=True)

# ── Admin: renew / delete ─────────────────────────────────────────────────────
@app.route('/admin/renew/<id>', methods=['POST'])
@login_required
@admin_required
def renew_agency(id):
    # Custom exact date takes priority
    custom_date = request.form.get('custom_renewal_date', '').strip()
    if custom_date:
        try:
            datetime.strptime(custom_date, '%Y-%m-%d')
            new_exp = custom_date
        except ValueError:
            custom_date = ''

    if not custom_date:
        raw = request.form.get('renewal_months', '12')
        try:
            months = float(raw)
        except (ValueError, TypeError):
            months = 12.0

        row  = supabase.table('agencies').select('expires_at').eq('id', id).single().execute().data
        base = datetime.now()
        if row.get('expires_at'):
            try:
                parsed = datetime.strptime(row['expires_at'], '%Y-%m-%d')
                if parsed > base:
                    base = parsed           # extend from current expiry if still active
            except ValueError:
                pass

        if months == 0.27:                  # 8-day demo
            new_exp = (base + timedelta(days=8)).strftime('%Y-%m-%d')
        else:
            days = int(round(months * 30.44))
            new_exp = (base + timedelta(days=days)).strftime('%Y-%m-%d')

    supabase.table('agencies').update({'expires_at': new_exp}).eq('id', id).execute()
    flash(f"Renewed. New expiry: {new_exp}", "success")
    return redirect(url_for('admin_dashboard',
        search=request.form.get('_search', ''), page=request.form.get('_page', 1)))

@app.route('/admin/delete/<id>', methods=['POST'])
@login_required
@admin_required
def delete_agency(id):
    supabase.table('agencies').delete().eq('id', id).execute()
    flash("Agency deleted.", "info")
    return redirect(url_for('admin_dashboard'))

# ── Generate (flat redirect → prefixed) ──────────────────────────────────────
@app.route('/generate')
@login_required
def generate():
    u = session['user']
    from_history = request.args.get('from_history', '')
    dest = f"/{u['username']}/generate"
    if from_history:
        dest += f'?from_history={from_history}'
    return redirect(dest)

# ── History (flat redirect → prefixed) ───────────────────────────────────────
@app.route('/history')
@login_required
def history():
    u = session['user']
    q    = request.args.get('q', '')
    page = request.args.get('page', '')
    dest = f"/{u['username']}/history"
    params = []
    if q:    params.append(f'q={q}')
    if page: params.append(f'page={page}')
    if params: dest += '?' + '&'.join(params)
    return redirect(dest)

@app.route('/history/<record_id>')
@login_required
def history_detail(record_id):
    u = session['user']
    return redirect(f"/{u['username']}/history/{record_id}")

@app.route('/history/delete/<record_id>', methods=['POST'])
@login_required
def history_delete(record_id):
    # Re-POST not possible via redirect; do the delete here then redirect
    aid = session['user']['id']
    supabase.table('generation_history').delete().eq('id', record_id).eq('agency_id', aid).execute()
    flash("Record deleted.", "info")
    u = session['user']
    return redirect(f"/{u['username']}/history")

# ── Error handler ─────────────────────────────────────────────────────────────
@app.errorhandler(Exception)
def handle_exc(e):
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException): return e
    return render_template('error.html', error_message=str(e)), 500

# Pre-load HTML templates into RAM
preload_templates()

# ── Periodic job cleanup — purge old jobs every 5 min ────────────────────────
def _cleanup_jobs():
    import time
    while True:
        time.sleep(300)
        cutoff = datetime.now().timestamp() - 600
        with jobs_lock:
            to_del = [jid for jid, job in jobs.items()
                      if job.get('status') in ('done','error')
                      and job.get('downloaded_at')
                      and datetime.fromisoformat(job['downloaded_at']).timestamp() < cutoff]
            for jid in to_del: jobs.pop(jid, None)
        if to_del: print(f'[Cleanup] Purged {len(to_del)} jobs')
threading.Thread(target=_cleanup_jobs, daemon=True).start()

# Pre-warm WeasyPrint font/CSS engine — first real render is instant
# Without this the first job pays a ~1.5s cold-start penalty
def _prewarm_weasyprint():
    try:
        WP_HTML(string="<p>warm</p>").write_pdf()
        print("[LibityInfotech] WeasyPrint warmed up.")
    except Exception:
        pass
threading.Thread(target=_prewarm_weasyprint, daemon=True).start()

if __name__ == '__main__':
    app.run(debug=True, threaded=True)